import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import os
from openpyxl import load_workbook

async def scrape_race_info(race_id: str):
    url = f"https://race.netkeiba.com/race/shutuba.html?race_id={race_id}"

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        await page.goto(url, timeout=60000)
        await page.wait_for_selector("tr.HorseList", timeout=30000)

        rows = await page.query_selector_all("tr.HorseList")

        horses = []
        for row in rows:
            try:
                uma_td = await row.query_selector("td[class^='Umaban']")
                uma_num = (await uma_td.inner_text()).strip() if uma_td else ""

                name_a = await row.query_selector("span.HorseName a")
                name = (await name_a.inner_text()).strip() if name_a else ""

                odds_span = await row.query_selector("td.Popular span")
                odds = (await odds_span.inner_text()).strip() if odds_span else ""

                horses.append({
                    "馬番": uma_num,
                    "馬名": name,
                    "単勝オッズ": odds
                })
            except Exception:
                continue

        await browser.close()

        horses.pop()
        horses.pop()

        # 出力ファイル名
        output_file = f"期待値計算_{race_id}.xlsx"

        if os.path.exists(output_file):
            # ▼ 既存ファイルの単勝オッズだけ更新（式は維持）
            wb = load_workbook(output_file)
            ws = wb.active

            for row in range(2, ws.max_row + 1):
                horse_name = ws[f"B{row}"].value
                for h in horses:
                    if h["馬名"] == horse_name:
                        ws[f"C{row}"] = h["単勝オッズ"]  # 単勝オッズのみ上書き

            wb.save(output_file)
            print(f"✅ オッズだけ更新（式は維持）: {output_file}")

        else:
            # ▼ 新規作成：数式も含めて書き込む
            df = pd.DataFrame(horses)
            df["勝率"] = [f"=1/{len(df)}*100" for i in range(len(df))]
            df["期待値"] = [f"=C{i+2}*D{i+2}" for i in range(len(df))]
            df["期待値順位"] = [f"=RANK(E{i+2},E$2:E${len(df)+1},0)" for i in range(len(df))]
            total_row = {
              "馬番": "",
              "馬名": "勝率合計",
              "単勝オッズ": "",
              "勝率": f"=SUM(D2:D{len(df)+1})",
              "期待値": "",
              "期待値順位": ""
            }
            df.loc[len(df)] = total_row  # 合計行を追加
            df.to_excel(output_file, index=False)
            print(f"✅ 新規作成: {output_file}")

# 使用例
race_id = "202501010411"
asyncio.run(scrape_race_info(race_id))
